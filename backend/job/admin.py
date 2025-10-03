# file: admin.py
from django.contrib import admin, messages
from django.http import HttpResponse
from django.db.models import Sum
import calendar
from datetime import datetime, date

import openpyxl

from .models import (
    Customer, JobDays, JobHours, HourlyRate, Advance, JobOvertimeHours, Dispute
)
from datetime import timedelta


@admin.action(description='Выгрузить в Excel за выбранные дни')
def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводный отчет"
    
    headers_summary = [
        'Сотрудник', 
        'Телефон',
        'Отработано дней',
        'Отработано часов',
        'Сверхурочные часы', 
        'Всего часов',
        'Часовая ставка',
        'Начислено зарплаты',
        'Выдано авансов',
        'Итого к выплате',
        'Баланс'
    ]
    
    for col, header in enumerate(headers_summary, 1):
        ws_summary.cell(row=1, column=col, value=header)
    
    selected_dates = [job_day.date for job_day in queryset]
    
    row = 2
    for customer in Customer.objects.all():
        # определяем ставку для сотрудника (индивидуальная или глобальная)
        rate = customer.hourly_rate if customer.hourly_rate is not None else HourlyRate.load().cash_per_hour

        normal_hours = JobHours.objects.filter(
            customer=customer, 
            date__in=queryset
        ).aggregate(total_hours=Sum('work_hours'))['total_hours'] or 0
        
        worked_days = JobHours.objects.filter(
            customer=customer, 
            date__in=queryset,
            work_start__isnull=False
        ).count()
        
        overtime_hours = JobOvertimeHours.objects.filter(
            customer=customer, 
            date__in=queryset,
            paid=True
        ).aggregate(total_hours=Sum('work_hours'))['total_hours'] or 0
        
        advances = Advance.objects.filter(
            customer=customer,
            date__date__in=selected_dates,
            accepted=True
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        total_hours = normal_hours + overtime_hours
        salary_earned = total_hours * rate
        total_to_pay = salary_earned - advances
        
        ws_summary.cell(row=row, column=1, value=customer.name)
        ws_summary.cell(row=row, column=2, value=customer.phone)
        ws_summary.cell(row=row, column=3, value=worked_days)
        ws_summary.cell(row=row, column=4, value=normal_hours)
        ws_summary.cell(row=row, column=5, value=overtime_hours)
        ws_summary.cell(row=row, column=6, value=total_hours)
        ws_summary.cell(row=row, column=7, value=rate)  # показываем индивидуальную ставку
        ws_summary.cell(row=row, column=8, value=salary_earned)
        ws_summary.cell(row=row, column=9, value=advances)
        ws_summary.cell(row=row, column=10, value=total_to_pay)
        ws_summary.cell(row=row, column=11, value=customer.balance)
        
        row += 1
    
    ws_daily = wb.create_sheet("Детализация по дням")
    
    headers_daily = [
        'Дата',
        'Сотрудник',
        'Приход',
        'Уход', 
        'Отработано часов',
        'Начало сверхурочки',
        'Конец сверхурочки',
        'Сверхурочные часы',
        'Статус оплаты',
        'Геолокация'
    ]
    
    for col, header in enumerate(headers_daily, 1):
        ws_daily.cell(row=1, column=col, value=header)
    
    row = 2
    for job_day in queryset:
        for job_hour in JobHours.objects.filter(date=job_day):
            ws_daily.cell(row=row, column=1, value=job_day.date.strftime('%d.%m.%Y'))
            ws_daily.cell(row=row, column=2, value=job_hour.customer.name if job_hour.customer else '')
            ws_daily.cell(row=row, column=3, value=job_hour.work_start.strftime('%H:%M') if job_hour.work_start else '')
            ws_daily.cell(row=row, column=4, value=job_hour.work_end.strftime('%H:%M') if job_hour.work_end else '')
            ws_daily.cell(row=row, column=5, value=job_hour.work_hours)
            ws_daily.cell(row=row, column=6, value='')
            ws_daily.cell(row=row, column=7, value='') 
            ws_daily.cell(row=row, column=8, value='')  
            ws_daily.cell(row=row, column=9, value='Оплачено' if job_hour.paid else 'Не оплачено')
            ws_daily.cell(row=row, column=10, value=job_hour.geolocation or '')
            row += 1
        
        for overtime in JobOvertimeHours.objects.filter(date=job_day):
            ws_daily.cell(row=row, column=1, value=job_day.date.strftime('%d.%m.%Y'))
            ws_daily.cell(row=row, column=2, value=overtime.customer.name if overtime.customer else '')
            ws_daily.cell(row=row, column=3, value='') 
            ws_daily.cell(row=row, column=4, value='') 
            ws_daily.cell(row=row, column=5, value='')  
            ws_daily.cell(row=row, column=6, value=overtime.work_start.strftime('%H:%M') if overtime.work_start else '')
            ws_daily.cell(row=row, column=7, value=overtime.work_end.strftime('%H:%M') if overtime.work_end else '')
            ws_daily.cell(row=row, column=8, value=overtime.work_hours)
            ws_daily.cell(row=row, column=9, value='Оплачено' if overtime.paid else 'Не оплачено')
            ws_daily.cell(row=row, column=10, value=overtime.geolocation or '')
            row += 1
    
    ws_financial = wb.create_sheet("Авансы и споры")
    
    headers_financial = [
        'Дата',
        'Тип',
        'Сотрудник', 
        'Сумма/Причина',
        'Статус'
    ]
    
    for col, header in enumerate(headers_financial, 1):
        ws_financial.cell(row=1, column=col, value=header)
    
    row = 2
    for advance in Advance.objects.filter(date__date__in=selected_dates):
        ws_financial.cell(row=row, column=1, value=advance.date.strftime('%d.%m.%Y %H:%M'))
        ws_financial.cell(row=row, column=2, value='Аванс')
        ws_financial.cell(row=row, column=3, value=advance.customer.name)
        ws_financial.cell(row=row, column=4, value=f"{advance.amount} сом - {advance.reason}")
        ws_financial.cell(row=row, column=5, value='Одобрен' if advance.accepted else 'На рассмотрении')
        row += 1
    
    for dispute in Dispute.objects.filter(date__date__in=selected_dates):
        ws_financial.cell(row=row, column=1, value=dispute.date.strftime('%d.%m.%Y %H:%M'))
        ws_financial.cell(row=row, column=2, value='Спор')
        ws_financial.cell(row=row, column=3, value=dispute.customer.name)
        ws_financial.cell(row=row, column=4, value=dispute.reason)
        ws_financial.cell(row=row, column=5, value='Решен' if dispute.resolved else 'Не решен')
        row += 1
    
    ws_stats = wb.create_sheet("Статистика по дням")
    
    headers_stats = [
        'Дата',
        'Всего сотрудников',
        'Пришедших',
        'Не пришедших',
        'Опоздавших',
        'Сверхурочных',
        'Среднее время работы'
    ]
    
    for col, header in enumerate(headers_stats, 1):
        ws_stats.cell(row=1, column=col, value=header)
    
    row = 2
    for job_day in queryset:
        total_employees = Customer.objects.count()
        came_employees = JobHours.objects.filter(date=job_day, work_start__isnull=False).count()
        not_came = total_employees - came_employees
        
        late_employees = JobHours.objects.filter(
            date=job_day, 
            work_start__gt=datetime.strptime('09:00', '%H:%M').time()
        ).count()
        
        overtime_count = JobOvertimeHours.objects.filter(date=job_day).count()
        
        sum_hours = JobHours.objects.filter(
            date=job_day
        ).aggregate(total=Sum('work_hours'))['total'] or 0
        avg_hours = (sum_hours / came_employees) if came_employees > 0 else 0
        
        ws_stats.cell(row=row, column=1, value=job_day.date.strftime('%d.%m.%Y'))
        ws_stats.cell(row=row, column=2, value=total_employees)
        ws_stats.cell(row=row, column=3, value=came_employees)
        ws_stats.cell(row=row, column=4, value=not_came)
        ws_stats.cell(row=row, column=5, value=late_employees)
        ws_stats.cell(row=row, column=6, value=overtime_count)
        ws_stats.cell(row=row, column=7, value=round(avg_hours, 2))
        row += 1
    
    # Стили и автоширина
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    if selected_dates:
        filename = f"отчет_{selected_dates[0].strftime('%d.%m.%Y')}_по_{selected_dates[-1].strftime('%d.%m.%Y')}.xlsx"
    else:
        filename = 'отчет.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@admin.action(description='Одобрить аванс у выбранных записей')
def accept_advance(modeladmin, request, queryset):
    for advance in queryset:
        advance.accepted = True
        # списываем с баланса сотрудника
        advance.customer.balance -= advance.amount
        advance.customer.save()
        advance.save()


@admin.action(description='Пополнить баланс по сверхурочным часам')
def replenish_balance_overtime(modeladmin, request, queryset):
    default_rate = HourlyRate.load().cash_per_hour
    for customer in queryset:
        rate = customer.hourly_rate if customer.hourly_rate is not None else default_rate
        for hours in customer.customer_overtime_hours.filter(paid=False):
            customer.balance += hours.work_hours * rate
            hours.paid = True
            hours.save()
        customer.save()


@admin.action(description='Пополнить баланс по отработанным часам')
def replenish_balance(modeladmin, request, queryset):
    default_rate = HourlyRate.load().cash_per_hour
    for customer in queryset:
        rate = customer.hourly_rate if customer.hourly_rate is not None else default_rate
        for hours in customer.customer_hours.filter(paid=False):
            customer.balance += hours.work_hours * rate
            hours.paid = True
            hours.save()
        customer.save()


@admin.action(description='Создать рабочие дни на следующие 30 дней (без выходных)')
def create_next_30_days(modeladmin, request, queryset):
    today = date.today()
    start_date = today + timedelta(days=1)  # начинаем с завтрашнего дня
    objs = []

    for i in range(30):
        d = start_date + timedelta(days=i)
        # Пропускаем субботу и воскресенье
        if d.weekday() >= 5:
            continue
        objs.append(JobDays(date=d))

    try:
        JobDays.objects.bulk_create(objs, ignore_conflicts=True)
        messages.success(request, f'Созданы/проверены рабочие дни на следующие 30 дней ({len(objs)} дат).')
    except TypeError:
        created = 0
        for obj in objs:
            obj, created_flag = JobDays.objects.get_or_create(date=obj.date)
            if created_flag:
                created += 1
        messages.success(request, f'Создано {created} новых рабочих дней на следующие 30 дней.')
    today = date.today()
    # первый день следующего месяца
    if today.month == 12:
        next_month_year = today.year + 1
        next_month = 1
    else:
        next_month_year = today.year
        next_month = today.month + 1

    _, days_in_month = calendar.monthrange(next_month_year, next_month)
    objs = []
    for day in range(1, days_in_month + 1):
        d = date(next_month_year, next_month, day)
        objs.append(JobDays(date=d))

    try:
        JobDays.objects.bulk_create(objs, ignore_conflicts=True)
        messages.success(request, f'Созданы/проверены дни на {next_month}.{next_month_year} ({len(objs)} дат).')
    except TypeError:
        # Для старых версий Django, где нет ignore_conflicts
        created = 0
        for obj in objs:
            obj, created_flag = JobDays.objects.get_or_create(date=obj.date)
            if created_flag:
                created += 1
        messages.success(request, f'Создано {created} новых дней на {next_month}.{next_month_year}.')


class JobHoursInline(admin.TabularInline):
    model = JobHours
    extra = 0 
    readonly_fields = ("work_hours", 'paid') 


class JobOvertimeHoursInline(admin.TabularInline):
    model = JobOvertimeHours
    extra = 0
    readonly_fields = ("work_hours", 'paid') 


@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    search_fields = ('name', 'phone')
    actions = [replenish_balance, replenish_balance_overtime]
    inlines = [JobHoursInline, JobOvertimeHoursInline]
    list_display = ("name", "phone", "hourly_rate", "balance")
    list_editable = ("hourly_rate",)


@admin.register(JobDays)
class JobDaysAdmin(admin.ModelAdmin):
    search_fields = ('date',)
    list_display = ("date",)
    inlines = [JobHoursInline, JobOvertimeHoursInline]
    actions = [export_to_excel, create_next_30_days]


@admin.register(JobHours)
class JobHoursAdmin(admin.ModelAdmin):
    list_display = ("customer", "date", "work_start", "work_end", "work_hours", "paid")
    list_filter = ("paid", "date__date", "customer__name")
    search_fields = ("customer__name", "date__date")
    readonly_fields = ("work_hours",)


@admin.register(JobOvertimeHours)
class JobOvertimeHoursAdmin(admin.ModelAdmin):
    list_display = ("customer", "date", "work_start", "work_end", "work_hours", "paid")
    list_filter = ("paid", "date__date", "customer__name")
    search_fields = ("customer__name", "date__date")
    readonly_fields = ("work_hours",)


@admin.register(Advance)
class AdvanceAdmin(admin.ModelAdmin):
    list_display = ("customer", "date", "amount", "reason")
    list_filter = ("date", "customer__name")
    search_fields = ("customer__name", "reason")
    actions = [accept_advance]


@admin.register(HourlyRate)
class HourlyRateAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        return not HourlyRate.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Dispute)
class DisputeAdmin(admin.ModelAdmin):
    list_display = ("customer", "date", "reason", "resolved")
    list_filter = ("resolved", "date", "customer__name")
    search_fields = ("customer__name", "reason")
